import os
import re
file_path = 'D:\\alt-text_report\\excel\\excel.xlsx'
#for excelfile in os.listdir(directory_path):
 #   if excelfile.endswith(".xlsx"):
        # Build the full file path
  #      file_path = os.path.join(directory_path, excelfile)
   #     print(file_path)

import openpyxl
workbook = openpyxl.load_workbook(file_path)
for sheet_name in workbook.sheetnames:
   sheet_obj = workbook[sheet_name]
   #print(f"Reading sheet: {sheet_name}")

   #print("Total Rows:", sheet_obj.max_row)
   Total_Rows = sheet_obj.max_row
   #print("Total Column:", sheet_obj.max_column)
   Total_Columns = sheet_obj.max_row
   max_row = sheet_obj.max_row
   cell1_hd1 = sheet_obj.cell(column = 7, row = 1)
   cell1_hd1_value = cell1_hd1.value
   cell1_hd9 = sheet_obj.cell(column = 9, row = 1)
   cell1_hd9_value = cell1_hd9.value
   cell1_hd2 = sheet_obj.cell(column = 10, row = 1)
   cell1_hd2_value = cell1_hd2.value
   cell1_hd3 = sheet_obj.cell(column = 11, row = 1)
   cell1_hd3_value = cell1_hd3.value
   cell1_hd4 = sheet_obj.cell(column = 12, row = 1)
   cell1_hd4_value = cell1_hd4.value

   #Row Heading validation
   if cell1_hd1_value == "Image File Name":
       #print("Valid heading")
      pass
   else:
        with open("D:\\alt-text_report\\Column_heading_invalid.htm", "a", encoding="utf-8") as f:
         f.write("Invalid Image File Name in " + str(sheet_name) + " - " + str(cell1_hd1_value) + str(cell1_hd1.coordinate)+ ": should be : should be Image File Name<br/>\n")

   if cell1_hd9_value == "Decorative":
       #print("Valid heading")
      pass
   else:
        with open("D:\\alt-text_report\\Column_heading_invalid.htm", "a", encoding="utf-8") as f:
         f.write("Invalid Decorative in " + str(sheet_name)  + " - " +  str(cell1_hd9_value)+ str(cell1_hd9.coordinate)+ ": should be Decorative<br/>\n")

   if cell1_hd2_value == "Alt text - Short":
       #print("Valid heading")
      pass
   else:
      with open("D:\\alt-text_report\\Column_heading_invalid.htm", "a", encoding="utf-8") as f:

         f.write("Invalid Alt text - Short in " + str(sheet_name)  + " - " +  str(cell1_hd2_value)+ str(cell1_hd2.coordinate)+ ": should be Alt text - Short<br/>\n")

   if cell1_hd3_value == "Extended Description Heading":
       #print("Valid heading")
      pass
   else:
        with open("D:\\alt-text_report\\Column_heading_invalid.htm", "a", encoding="utf-8") as f:
         f.write("Invalid Extended Alt text in " + str(sheet_name)  + " - " +  str(cell1_hd3_value)+ str(cell1_hd3.coordinate)+ ": should be Extended Description Heading<br/>\n")

   if cell1_hd4_value == "Extended Alt text":
       #print("Valid heading")
       with open("D:\\alt-text_report\\excel_alt_text.htm", "a", encoding="utf-8") as f:
        f.write('''<table border="1"><br/>\n<tr><th>Image_File_Name</th><th>Thumbnail of Image</th><th>Decorative</th><th>Alt_text_Short</th><th>Extended_Alt_text</th><th>Extended_Description_Heading</th></tr>''')

   else:
        with open("D:\\alt-text_report\\Column_heading_invalid.htm", "a", encoding="utf-8") as f:
         f.write("Invalid Extended Description Heading in " + str(sheet_name)  + " - " +  str(cell1_hd4_value)+ str(cell1_hd4.coordinate)+ ": should be <br/>\n")


   for i in range(2, max_row + 1):
       cell1_obj = sheet_obj.cell(column = 7, row = i)
       cell_image = sheet_obj.cell(column = 8, row = i)

       cell9_obj = sheet_obj.cell(column = 9, row = i)
       cell2_obj = sheet_obj.cell(column = 10, row = i)
       cell3_obj = sheet_obj.cell(column = 11, row = i)
       cell4_obj = sheet_obj.cell(column = 12, row = i)
       cell_image_name = cell_image.coordinate
       sheet_name = sheet_obj.title

       Image_File_Name = cell1_obj.value
       decorative = cell9_obj.value
       Alt_text_Short = cell2_obj.value
       Extended_Alt_text = cell4_obj.value
       Extended_Description_Heading = cell3_obj.value
       with open("D:\\alt-text_report\\excel_alt_text.htm", "a", encoding="utf-8") as f:
#             f.write(str(sheet_name) + "\t" + str(row_col_name) + "\t" + str(Image_File_Name) + "\n")
        f.write("<tr><td>" + str(Image_File_Name) +
            '''</td><td><img width="350px" src="Excel_images/image_''' + str(sheet_name) +"_" + str(cell_image_name) +  
            '''.png"/></td><td>''' + str(decorative) +
            "</td><td>" + str(Alt_text_Short) +
            "</td><td>" + str(Extended_Alt_text) +
            "</td><td>" + str(Extended_Description_Heading) + "</td></tr>"
            "\n")

with open("D:\\alt-text_report\\excel_alt_text.htm", "a", encoding="utf-8") as f:
 f.write("</table>")
print("Excel_Finished")



html_directory_path = 'D:\\alt-text_report\\xhtml\\'
import chardet
with open("D:\\alt-text_report\\html_alttext.htm", "a", encoding="utf-8") as f:
 f.write('''<table border="1">\n<tr><th>Image_File_Name</th><th>Thumbnail of Image</th><th>Decorative</th><th>Alt_text_Short</th><th>Extended_Alt_text</th><th>Extended_Description_Heading</th></tr>\n''')
for xhtmlfile in os.listdir(html_directory_path):
    if xhtmlfile.endswith(".xhtml"):
        # Build the full file path
        file_path = os.path.join(html_directory_path, xhtmlfile)
    with open(file_path, encoding="utf-8", errors='replace') as f:
        htmlfile = f.read()
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(htmlfile, 'html.parser')

    images_not_in_annotation = soup.select('body img:not(annotation-xml img)')
    for img in images_not_in_annotation:
        if img and 'aria-details' in img.attrs:
            aria_details_id = img['aria-details']
            aria_details_ref = img.get('aria-details')
            element_details = soup.find(id=aria_details_ref)
            #print(element_details)
            summary = element_details.find('summary')
            #print(summary.sourceline, xhtmlfile, summary)
            summary_text = summary.get_text()
            cleaned_summary = summary_text.replace("Extended description for", '', 1)
            imgsrc = img['src']
            img_alt = img['alt']
            
            cleaned_src = imgsrc.replace("../images/", '', 1)
            
            with open("D:\\alt-text_report\\html_alttext.htm", "a", encoding="utf-8") as f:
             f.write('''<tr class="td" style="background-color: gray;"><td>'''+str(cleaned_src) + '''</td><td><img width="350px" src="HTML_images/''' + str(imgsrc) +'''"/></td><td></td><td>'''+ str(img_alt) + '''</td><td>''' + str(element_details) + "</td><td>" + str(cleaned_summary) + "</td></tr>" "\n")
            with open("D:\\alt-text_report\\images_name.htm", "a", encoding="utf-8") as f:
             f.write(str(cleaned_src)+"\n")
        else:
           try:
            imgsrc = img['src']
            img_alt = img['alt']
            #print(img_alt)
            
                  
            cleaned_src = imgsrc.replace("../images/", '', 1)
               
            with open("D:\\alt-text_report\\html_alttext.htm", "a", encoding="utf-8") as f:
                f.write('''<tr class="td" style="background-color: gray;"><td>'''+str(cleaned_src) + '''</td><td><img width="351px" src="HTML_images/''' + str(imgsrc) +'''"/></td><td></td><td>'''+ str(img_alt) + "</td><td></td><td></td></tr>" "\n")
            with open("D:\\alt-text_report\\images_name.htm", "a", encoding="utf-8") as f:
                f.write(str(cleaned_src)+"\n")
           except Exception as e:
            with open("D:\\alt-text_report\\Error.htm", "a", encoding="utf-8") as f:
                f.write(f"{imgsrc}\t{e} attribute missing\n<br></br>")

with open("D:\\alt-text_report\\html_alttext.htm", "a", encoding="utf-8") as f:
 f.write("</table>")

print('Finished')


with open("D:\\alt-text_report\\Extract_excel_html_alt-text_finished.txt", "a", encoding="utf-8") as f:
    f.write("task completed")
print('Finished')
